"""apps/reports/views.py – Full report suite with working Excel & PDF export.
Fixes:
  1. build_excel() and build_pdf() were missing — ALL downloads failed.
  2. DashboardSummaryView now returns fuel_litres, expenditure, trips,
     fuel_excess_events, stock_value, stock_items, and expiry_alerts.
  3. FIXED: __import__('django') hack replaced with proper Q import.
"""
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Sum, Count, Q          # ✅ FIX: Added Q here
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_date_range(request, default_days=30):
    today = date.today()
    date_from = request.query_params.get('date_from') or str(today - timedelta(days=default_days))
    date_to   = request.query_params.get('date_to')   or str(today)
    return date_from, date_to


def _excel_response(filename):
    r = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _pdf_response(filename):
    r = HttpResponse(content_type='application/pdf')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


# ── Excel builder (openpyxl) ──────────────────────────────────────────────────

def build_excel(title, headers, rows, summary=None):
    """Return a BytesIO of an .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # ── Title row ──
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1a3a6e")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row ──
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(border_style="thin", color="D1D5DB")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    # ── Data rows ──
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for row_idx, row in enumerate(rows, 3):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if isinstance(val, (int, float)) and col_idx > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # ── Summary block ──
    if summary:
        gap_row = len(rows) + 4
        ws.cell(row=gap_row, column=1, value="SUMMARY").font = Font(bold=True, size=11)
        for i, (k, v) in enumerate(summary.items()):
            r = gap_row + 1 + i
            ws.cell(row=r, column=1, value=k).font = Font(bold=True)
            cell = ws.cell(row=r, column=2, value=v)
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0.00'

    # ── Auto column widths ──
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in rows:
            try:
                max_len = max(max_len, len(str(row[col_idx - 1])))
            except IndexError:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF builder (reportlab) ───────────────────────────────────────────────────

def build_pdf(title, headers, rows, summary=None):
    """Return a BytesIO of a PDF file."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise ImportError("reportlab is required: pip install reportlab")

    buf = BytesIO()
    page = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    brand  = colors.HexColor('#1a3a6e')
    accent = colors.HexColor('#2563EB')

    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Title'],
        fontSize=16, textColor=brand, spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        'SubStyle', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, spaceAfter=12,
    )

    story = [
        Paragraph("Taurus Trade & Logistics ERP", sub_style),
        Paragraph(title, title_style),
        Paragraph(f"Generated: {date.today().strftime('%d %B %Y')}", sub_style),
        Spacer(1, 0.3*cm),
    ]

    # ── Data table ──
    col_w = (page[0] - 3*cm) / len(headers)
    table_data = [headers] + [list(map(str, r)) for r in rows]
    tbl = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), accent),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 9),
        ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('TOPPADDING',   (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl)

    # ── Summary table ──
    if summary:
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Summary", ParagraphStyle(
            'SumHead', parent=styles['Heading2'], fontSize=11, textColor=brand,
        )))
        sum_data = [[str(k), f"GH₵ {float(v):,.2f}" if isinstance(v, (int, float)) else str(v)]
                    for k, v in summary.items()]
        stbl = Table(sum_data, colWidths=[8*cm, 5*cm])
        stbl.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (0, -1), colors.HexColor('#EFF6FF')),
            ('FONTNAME',     (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1), 9),
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#BFDBFE')),
            ('TOPPADDING',   (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
            ('LEFTPADDING',  (0, 0), (-1, -1), 8),
        ]))
        story.append(stbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ================================================================
# REPORT VIEWS
# ================================================================

class RevenueExpenditureReportView(APIView):
    """Revenue vs Expenditure P&L report — all sources."""
    def get(self, request):
        from apps.finance.models import Revenue, Expenditure
        date_from, date_to = _parse_date_range(request)

        rev_qs = (
            Revenue.objects
            .filter(date__range=[date_from, date_to])
            .values('source')
            .annotate(total=Sum('amount'))
            .order_by('-total')
        )
        exp_qs = (
            Expenditure.objects
            .filter(date__range=[date_from, date_to])
            .values('category')
            .annotate(total=Sum('amount'))
            .order_by('-total')
        )

        total_rev = sum(r['total'] or 0 for r in rev_qs)
        total_exp = sum(e['total'] or 0 for e in exp_qs)
        net = total_rev - total_exp

        revenue_by_source = {r['source']: float(r['total'] or 0) for r in rev_qs}

        headers = ['Category / Source', 'Type', 'Amount (GH₵)']
        rows = []
        for r in rev_qs:
            rows.append([r['source'].replace('_', ' '), 'REVENUE', float(r['total'] or 0)])
        for e in exp_qs:
            rows.append([e['category'].replace('_', ' '), 'EXPENDITURE', float(e['total'] or 0)])

        summary = {
            'Total Revenue (GH₵)':        float(total_rev),
            'Total Expenditure (GH₵)':    float(total_exp),
            'Net Profit / Loss (GH₵)':    float(net),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Revenue vs Expenditure  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'pnl_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'pnl_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary,
                         'revenue_by_source': revenue_by_source})


class FuelReportView(APIView):
    """Fuel consumption report with excess incidents."""
    def get(self, request):
        from apps.fuel.models import FuelLog
        date_from, date_to = _parse_date_range(request)

        logs = (
            FuelLog.objects
            .select_related('truck', 'trip')
            .filter(date__range=[date_from, date_to])
            .order_by('-date')
        )

        headers = ['Date', 'Truck', 'Trip', 'Litres', 'Limit (L)',
                   'Excess (L)', 'Price/L', 'Total Cost', 'Excess Cost', 'Remark']
        rows = []
        total_litres = Decimal('0')
        total_cost   = Decimal('0')
        total_excess = Decimal('0')
        excess_events = 0

        for log in logs:
            excess_cost = log.excess_fuel * log.price_per_litre
            rows.append([
                str(log.date),
                log.truck.truck_number,
                log.trip.waybill_no if log.trip else '—',
                float(log.litres),
                float(log.fuel_limit),
                float(log.excess_fuel),
                float(log.price_per_litre),
                float(log.total_cost),
                float(excess_cost),
                log.remark or '—',
            ])
            total_litres += log.litres
            total_cost   += log.total_cost
            total_excess += log.excess_fuel
            if log.excess_fuel > 0:
                excess_events += 1

        summary = {
            'Total Litres Issued':      float(total_litres),
            'Total Fuel Cost (GH₵)':   float(total_cost),
            'Total Excess Litres':      float(total_excess),
            'Excess Incidents':         excess_events,
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Fuel Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'fuel_report_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'fuel_report_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class TripReportView(APIView):
    """Trip report with revenue, qty and delivery data."""
    def get(self, request):
        from apps.trips.models import Trip
        date_from, date_to = _parse_date_range(request)

        trips = (
            Trip.objects
            .select_related('truck', 'driver')
            .filter(loading_time__date__range=[date_from, date_to])
            .order_by('-loading_time')
        )

        headers = ['Waybill', 'Date', 'Truck', 'Driver', 'Origin', 'Destination',
                   'Material', 'Loaded (t)', 'Delivered (t)', 'Diff (t)', 'Revenue (GH₵)', 'Status']
        rows = []
        total_revenue = Decimal('0')
        total_loaded  = Decimal('0')

        for t in trips:
            rows.append([
                t.waybill_no,
                str(t.loading_time.date()),
                t.truck.truck_number,
                f"{t.driver.first_name} {t.driver.last_name}",
                t.origin,
                t.destination,
                t.material_type,
                float(t.loaded_qty),
                float(t.delivered_qty or 0),
                float(t.qty_difference or 0),
                float(t.trip_revenue),
                t.status,
            ])
            total_revenue += t.trip_revenue
            total_loaded  += t.loaded_qty

        summary = {
            'Total Trips':              len(rows),
            'Total Loaded (tonnes)':    float(total_loaded),
            'Total Revenue (GH₵)':     float(total_revenue),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Trip Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'trips_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'trips_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class StockReportView(APIView):
    """Current stock levels and valuations."""
    def get(self, request):
        from apps.inventory.models import Item, StockLedger

        items = Item.objects.filter(is_deleted=False).order_by('item_type', 'name')

        headers = ['Item', 'Type', 'Unit', 'Qty in Stock', 'Stock Value (GH₵)', 'Reorder Level', 'Status']
        rows = []
        total_value = Decimal('0')

        for item in items:
            qty   = item.available_qty()
            value = item.available_value()
            total_value += value
            status = 'LOW STOCK' if qty <= item.reorder_level and item.reorder_level > 0 else 'OK'
            rows.append([
                item.name,
                item.get_item_type_display(),
                item.unit,
                float(qty),
                float(value),
                float(item.reorder_level),
                status,
            ])

        summary = {
            'Total Items':              len(rows),
            'Total Stock Value (GH₵)': float(total_value),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Stock Report — as of {date.today()}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'stock_{date.today()}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'stock_{date.today()}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class InvoiceReportView(APIView):
    """Invoice listing with VAT and payment status."""
    def get(self, request):
        from apps.invoicing.models import Invoice
        date_from, date_to = _parse_date_range(request)

        invoices = (
            Invoice.objects
            .prefetch_related('lines')
            .select_related('trip')
            .filter(invoice_date__range=[date_from, date_to])
            .order_by('-invoice_date')
        )

        headers = ['Invoice #', 'Date', 'Client', 'Trip', 'Qty', 'Subtotal', 'VAT', 'Total', 'Status']
        rows = []
        total_invoiced = Decimal('0')
        total_paid     = Decimal('0')

        for inv in invoices:
            qty = sum(l.quantity for l in inv.lines.all())
            rows.append([
                inv.invoice_number,
                str(inv.invoice_date),
                inv.client_name,
                inv.trip.waybill_no if inv.trip else '—',
                float(qty),
                float(inv.subtotal),
                float(inv.vat_amount),
                float(inv.total_amount),
                inv.status,
            ])
            total_invoiced += inv.total_amount
            if inv.status == 'PAID':
                total_paid += inv.total_amount

        summary = {
            'Total Invoiced (GH₵)':  float(total_invoiced),
            'Received (GH₵)':        float(total_paid),
            'Outstanding (GH₵)':     float(total_invoiced - total_paid),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Invoice Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'invoices_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'invoices_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class SparePartsReportView(APIView):
    """Purchases and issues of spare parts."""
    def get(self, request):
        from apps.inventory.models import StockLedger
        date_from, date_to = _parse_date_range(request)

        ledger = (
            StockLedger.objects
            .select_related('item', 'location')
            .filter(
                item__item_type='SPARE_PART',
                created_at__date__range=[date_from, date_to],
            )
            .order_by('-created_at')
        )

        headers = ['Date', 'Item', 'Transaction', 'Qty', 'Unit Cost', 'Amount (GH₵)', 'Location', 'Reference']
        rows = []
        total_in  = Decimal('0')
        total_out = Decimal('0')

        for entry in ledger:
            qty = float(entry.quantity)
            rows.append([
                str(entry.created_at.date()),
                entry.item.name,
                entry.get_transaction_type_display(),
                qty,
                float(entry.unit_cost or 0),
                float(entry.final_amount or 0),
                entry.location.name if entry.location else '—',
                entry.reference or '—',
            ])
            if entry.quantity > 0:
                total_in += entry.final_amount or 0
            else:
                total_out += abs(entry.final_amount or 0)

        summary = {
            'Total Inward Value (GH₵)':  float(total_in),
            'Total Issued Value (GH₵)':  float(total_out),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Spare Parts Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'spare_parts_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'spare_parts_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class MaintenanceReportView(APIView):
    """Service history and maintenance costs."""
    def get(self, request):
        from apps.maintenance.models import MaintenanceRecord
        date_from, date_to = _parse_date_range(request)

        records = (
            MaintenanceRecord.objects
            .select_related('truck')
            .filter(date__range=[date_from, date_to])
            .order_by('-date')
        )

        headers = ['Date', 'Truck', 'Type', 'Description', 'Cost (GH₵)', 'Vendor', 'Next Due']
        rows = []
        total_cost = Decimal('0')

        for rec in records:
            rows.append([
                str(rec.date),
                rec.truck.truck_number,
                rec.maintenance_type if hasattr(rec, 'maintenance_type') else '—',
                rec.description[:60] if rec.description else '—',
                float(rec.cost or 0),
                rec.vendor.name if hasattr(rec, 'vendor') and rec.vendor else '—',
                str(rec.next_due_date) if hasattr(rec, 'next_due_date') and rec.next_due_date else '—',
            ])
            total_cost += rec.cost or 0

        summary = {'Total Maintenance Cost (GH₵)': float(total_cost)}

        fmt = request.query_params.get('format', 'json')
        report_title = f'Maintenance Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'maintenance_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'maintenance_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class VATReportView(APIView):
    """VAT charged on invoices."""
    def get(self, request):
        from apps.invoicing.models import Invoice
        date_from, date_to = _parse_date_range(request)

        invoices = (
            Invoice.objects
            .filter(invoice_date__range=[date_from, date_to], vat_enabled=True)
            .order_by('-invoice_date')
        )

        headers = ['Invoice #', 'Date', 'Client', 'Subtotal (GH₵)', 'VAT %', 'VAT Amount (GH₵)', 'Total (GH₵)', 'Status']
        rows = []
        total_vat = Decimal('0')

        for inv in invoices:
            rows.append([
                inv.invoice_number,
                str(inv.invoice_date),
                inv.client_name,
                float(inv.subtotal),
                float(inv.vat_percent or 0),
                float(inv.vat_amount),
                float(inv.total_amount),
                inv.status,
            ])
            total_vat += inv.vat_amount

        summary = {'Total VAT Collected (GH₵)': float(total_vat)}

        fmt = request.query_params.get('format', 'json')
        report_title = f'VAT Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'vat_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'vat_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class TyreReportView(APIView):
    """Tyre inventory, fitment and wear status."""
    def get(self, request):
        from apps.inventory.models import StockLedger
        date_from, date_to = _parse_date_range(request)

        ledger = (
            StockLedger.objects
            .select_related('item', 'location')
            .filter(
                item__item_type='TYRE',
                created_at__date__range=[date_from, date_to],
            )
            .order_by('-created_at')
        )

        headers = ['Date', 'Item', 'Transaction', 'Qty', 'Unit Cost', 'Amount (GH₵)', 'Location']
        rows = []
        total_value = Decimal('0')

        for entry in ledger:
            rows.append([
                str(entry.created_at.date()),
                entry.item.name,
                entry.get_transaction_type_display(),
                float(entry.quantity),
                float(entry.unit_cost or 0),
                float(entry.final_amount or 0),
                entry.location.name if entry.location else '—',
            ])
            total_value += abs(entry.final_amount or 0)

        summary = {'Total Tyre Value (GH₵)': float(total_value)}

        fmt = request.query_params.get('format', 'json')
        report_title = f'Tyre Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'tyres_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'tyres_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


class LubricantReportView(APIView):
    """Lubricant purchases, issues and consumption."""
    def get(self, request):
        from apps.inventory.models import StockLedger
        date_from, date_to = _parse_date_range(request)

        ledger = (
            StockLedger.objects
            .select_related('item', 'location')
            .filter(
                item__item_type='LUBRICANT',
                created_at__date__range=[date_from, date_to],
            )
            .order_by('-created_at')
        )

        headers = ['Date', 'Item', 'Transaction', 'Qty', 'Unit', 'Unit Cost (GH₵)', 'Amount (GH₵)', 'Location', 'Reference']
        rows = []
        total_in  = Decimal('0')
        total_out = Decimal('0')

        for entry in ledger:
            rows.append([
                str(entry.created_at.date()),
                entry.item.name,
                entry.get_transaction_type_display(),
                float(entry.quantity),
                entry.item.unit,
                float(entry.unit_cost or 0),
                float(entry.final_amount or 0),
                entry.location.name if entry.location else '—',
                entry.reference or '—',
            ])
            if entry.quantity > 0:
                total_in  += entry.final_amount or 0
            else:
                total_out += abs(entry.final_amount or 0)

        summary = {
            'Total Purchased Value (GH₵)': float(total_in),
            'Total Issued Value (GH₵)':    float(total_out),
        }

        fmt = request.query_params.get('format', 'json')
        report_title = f'Lubricant Report  {date_from} → {date_to}'

        if fmt == 'excel':
            buf = build_excel(report_title, headers, rows, summary)
            resp = _excel_response(f'lubricants_{date_from}_{date_to}.xlsx')
            resp.write(buf.read())
            return resp
        if fmt == 'pdf':
            buf = build_pdf(report_title, headers, rows, summary)
            resp = _pdf_response(f'lubricants_{date_from}_{date_to}.pdf')
            resp.write(buf.read())
            return resp

        return Response({'headers': headers, 'rows': rows, 'summary': summary})


# ================================================================
# DASHBOARD — FIXED (was missing fuel, trips, stock, alerts)
# ================================================================

class DashboardSummaryView(APIView):
    """Complete KPI summary for dashboard."""
    def get(self, request):
        from apps.trucks.models import Truck
        from apps.drivers.models import Driver
        from apps.trips.models import Trip
        from apps.finance.models import Revenue, Expenditure
        from apps.fuel.models import FuelLog
        from apps.inventory.models import Item, StockLedger
        from django.utils import timezone

        today       = timezone.now().date()
        month_start = today.replace(day=1)

        # ── Fuel this month ── ✅ FIXED: Q imported at top, no more __import__ hack
        fuel_agg = (
            FuelLog.objects
            .filter(date__gte=month_start)
            .aggregate(
                litres=Sum('litres'),
                excess_events=Count('id', filter=Q(excess_fuel__gt=0))
            )
        )

        # ── Stock value ──
        stock_value_agg   = StockLedger.objects.aggregate(val=Sum('final_amount'))
        stock_items_count = Item.objects.filter(is_deleted=False).count()

        # ── Expiry alerts (next 30 days) ──
        trucks = Truck.objects.filter(status='ACTIVE')
        expiry_alerts = []
        for truck in trucks:
            for alert in truck.expiry_alerts():
                expiry_alerts.append(alert)
        expiry_alerts.sort(key=lambda a: a['days_remaining'])

        return Response({
            'fleet': {
                'active_trucks':  Truck.objects.filter(status='ACTIVE').count(),
                'active_drivers': Driver.objects.filter(status='ACTIVE').count(),
                'ongoing_trips':  Trip.objects.filter(status='EN_ROUTE').count(),
            },
            'this_month': {
                'revenue':            float(Revenue.objects.filter(date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0),
                'expenditure':        float(Expenditure.objects.filter(date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0),
                'trips':              Trip.objects.filter(loading_time__date__gte=month_start).count(),
                'fuel_litres':        float(fuel_agg.get('litres') or 0),
                'fuel_excess_events': fuel_agg.get('excess_events') or 0,
            },
            'stock_value': float(stock_value_agg.get('val') or 0),
            'stock_items': stock_items_count,
            'expiry_alerts': expiry_alerts,
        })
